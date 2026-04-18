#!/usr/bin/env python3
"""Generate categorized expense report Excel from invoice data (JSON input)."""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

CATEGORIES = {
    "业务招待": ["客户餐费", "商务宴请", "礼品", "娱乐", "招待"],
    "差旅-交通": ["机票", "火车票", "高铁", "出租车", "打车", "地铁", "公交", "网约车", "加油", "过路费", "停车费"],
    "差旅-住宿": ["酒店", "住宿", "宾馆"],
    "差旅-餐饮": ["出差餐费", "差旅餐饮", "工作餐"],
    "办公用品": ["文具", "打印", "办公耗材", "办公设备"],
    "通讯费": ["话费", "流量", "网费", "宽带"],
    "其他": [],
}

HEADER_FILL = PatternFill(start_color="CC785C", end_color="CC785C", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
CATEGORY_FILL = PatternFill(start_color="F5E6DE", end_color="F5E6DE", fill_type="solid")
CATEGORY_FONT = Font(name="Microsoft YaHei", bold=True, size=11)
BODY_FONT = Font(name="Microsoft YaHei", size=10)
MONEY_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

COLUMNS = ["日期", "商户/来源", "项目", "金额(元)", "备注"]
COL_WIDTHS = [14, 22, 16, 14, 28]


def style_header(ws, row):
    for col_idx, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_category_row(ws, row, name, subtotal, count):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value=f"{name}（{count}笔）")
    cell.font = CATEGORY_FONT
    cell.fill = CATEGORY_FILL
    cell.border = THIN_BORDER
    amount_cell = ws.cell(row=row, column=4, value=subtotal)
    amount_cell.font = CATEGORY_FONT
    amount_cell.fill = CATEGORY_FILL
    amount_cell.number_format = MONEY_FORMAT
    amount_cell.border = THIN_BORDER
    for c in [2, 3, 5]:
        ws.cell(row=row, column=c).fill = CATEGORY_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER


def write_item(ws, row, item):
    values = [
        item.get("date", ""),
        item.get("vendor", ""),
        item.get("item", ""),
        item.get("amount", 0),
        item.get("note", ""),
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        if col_idx == 4:
            cell.number_format = MONEY_FORMAT
            cell.alignment = Alignment(horizontal='right')


def generate(invoices, output_path):
    # Group by category
    grouped = {cat: [] for cat in CATEGORIES}
    for inv in invoices:
        cat = inv.get("category", "其他")
        if cat not in grouped:
            cat = "其他"
        grouped[cat].append(inv)

    wb = Workbook()
    ws = wb.active
    ws.title = "发票报销汇总"

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value=f"发票报销汇总表")
    title_cell.font = Font(name="Microsoft YaHei", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:E2')
    date_cell = ws.cell(row=2, column=1, value=f"生成日期: {datetime.now().strftime('%Y-%m-%d')}")
    date_cell.font = Font(name="Microsoft YaHei", size=9, color="888888")
    date_cell.alignment = Alignment(horizontal='right')

    row = 4
    style_header(ws, row)
    row += 1

    grand_total = 0

    for cat_name in CATEGORIES:
        items = grouped[cat_name]
        if not items:
            continue
        subtotal = sum(i.get("amount", 0) for i in items)
        grand_total += subtotal

        style_category_row(ws, row, cat_name, subtotal, len(items))
        row += 1

        # Sort items by date
        items.sort(key=lambda x: x.get("date", ""))
        for item in items:
            write_item(ws, row, item)
            row += 1

    # Grand total row
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    total_label = ws.cell(row=row, column=1, value="合计")
    total_label.font = Font(name="Microsoft YaHei", bold=True, size=12)
    total_label.alignment = Alignment(horizontal='right')
    total_label.border = THIN_BORDER
    total_cell = ws.cell(row=row, column=4, value=grand_total)
    total_cell.font = Font(name="Microsoft YaHei", bold=True, size=12)
    total_cell.number_format = MONEY_FORMAT
    total_cell.border = THIN_BORDER
    for c in [2, 3, 5]:
        ws.cell(row=row, column=c).border = THIN_BORDER

    # Also create a per-category summary sheet
    ws2 = wb.create_sheet("分类汇总")
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 14
    ws2.cell(row=1, column=1, value="报销类别").font = HEADER_FONT
    ws2.cell(row=1, column=1).fill = HEADER_FILL
    ws2.cell(row=1, column=1).border = THIN_BORDER
    ws2.cell(row=1, column=2, value="笔数").font = HEADER_FONT
    ws2.cell(row=1, column=2).fill = HEADER_FILL
    ws2.cell(row=1, column=2).border = THIN_BORDER
    ws2.cell(row=1, column=3, value="小计(元)").font = HEADER_FONT
    ws2.cell(row=1, column=3).fill = HEADER_FILL
    ws2.cell(row=1, column=3).border = THIN_BORDER

    s_row = 2
    for cat_name in CATEGORIES:
        items = grouped[cat_name]
        if not items:
            continue
        subtotal = sum(i.get("amount", 0) for i in items)
        ws2.cell(row=s_row, column=1, value=cat_name).font = BODY_FONT
        ws2.cell(row=s_row, column=1).border = THIN_BORDER
        ws2.cell(row=s_row, column=2, value=len(items)).font = BODY_FONT
        ws2.cell(row=s_row, column=2).border = THIN_BORDER
        c = ws2.cell(row=s_row, column=3, value=subtotal)
        c.font = BODY_FONT
        c.number_format = MONEY_FORMAT
        c.border = THIN_BORDER
        s_row += 1

    ws2.cell(row=s_row, column=1, value="合计").font = CATEGORY_FONT
    ws2.cell(row=s_row, column=1).border = THIN_BORDER
    ws2.cell(row=s_row, column=2, value=sum(len(grouped[c]) for c in CATEGORIES)).font = CATEGORY_FONT
    ws2.cell(row=s_row, column=2).border = THIN_BORDER
    tc = ws2.cell(row=s_row, column=3, value=grand_total)
    tc.font = CATEGORY_FONT
    tc.number_format = MONEY_FORMAT
    tc.border = THIN_BORDER

    wb.save(output_path)
    return output_path, len(invoices), grand_total


def main():
    parser = argparse.ArgumentParser(description="Generate categorized expense report Excel")
    parser.add_argument("--input", "-i", required=True, help="JSON file with invoice data")
    parser.add_argument("--output", "-o", help="Output .xlsx path (default: 发票报销汇总-{date}.xlsx)")
    args = parser.parse_args()

    with open(args.input, 'r', encoding='utf-8') as f:
        data = json.load(f)

    invoices = data if isinstance(data, list) else data.get("invoices", [])

    if not args.output:
        date_str = datetime.now().strftime('%Y%m%d')
        args.output = f"发票报销汇总-{date_str}.xlsx"

    out_path, count, total = generate(invoices, args.output)
    print(f"Generated: {out_path}")
    print(f"  Invoices: {count}")
    print(f"  Total: ¥{total:,.2f}")


if __name__ == "__main__":
    main()
